"""Safe, bounded ingestion endpoints for tabular audit data."""

from __future__ import annotations

import csv
import io
import math
import os
import sqlite3
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

from flask import Blueprint, current_app, jsonify, request
from openpyxl import load_workbook

from models import AuditArea, DataSource, db, utcnow


data_sources_bp = Blueprint("data_source_ingestion", __name__)
ALLOWED_TABULAR_EXTENSIONS = {".csv", ".xlsx"}
DEFAULT_MAX_BYTES = 10 * 1024 * 1024
DEFAULT_MAX_ROWS = 10_000
DEFAULT_MAX_COLUMNS = 100


class IngestionError(ValueError):
    """A user-correctable source validation error."""


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, bool, int)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, (bytes, bytearray)):
        return value.hex()
    return str(value)


def _unique_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, raw in enumerate(values, start=1):
        base = str(raw).strip() if raw is not None else ""
        base = base or f"column_{index}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers


def _column_metadata(headers: list[str], records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for header in headers:
        values = [row.get(header) for row in records if row.get(header) not in (None, "")]
        types = sorted({type(value).__name__ for value in values})
        result.append({"name": header, "types": types or ["null"], "nullable": len(values) < len(records)})
    return result


def _bounded_rows(rows, headers: list[str], max_rows: int) -> list[dict[str, Any]]:
    records = []
    for row_number, row in enumerate(rows, start=1):
        if row_number > max_rows:
            raise IngestionError(f"source exceeds the {max_rows} row limit")
        values = list(row)
        records.append({header: _json_value(values[index] if index < len(values) else None)
                        for index, header in enumerate(headers)})
    return records


def parse_csv(content: bytes, max_rows: int, max_columns: int) -> tuple[list[dict], list[dict]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise IngestionError("CSV files must use UTF-8 encoding") from exc
    reader = csv.reader(io.StringIO(text))
    try:
        headers = _unique_headers(next(reader))
    except StopIteration as exc:
        raise IngestionError("CSV file is empty") from exc
    if len(headers) > max_columns:
        raise IngestionError(f"source exceeds the {max_columns} column limit")
    records = _bounded_rows(reader, headers, max_rows)
    return records, _column_metadata(headers, records)


def parse_xlsx(content: bytes, max_rows: int, max_columns: int, sheet_name: str | None = None):
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise IngestionError("invalid or unsupported XLSX file") from exc
    try:
        if sheet_name and sheet_name not in workbook.sheetnames:
            raise IngestionError(f"worksheet not found: {sheet_name}")
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = _unique_headers(list(next(rows)))
        except StopIteration as exc:
            raise IngestionError("worksheet is empty") from exc
        if len(headers) > max_columns:
            raise IngestionError(f"source exceeds the {max_columns} column limit")
        records = _bounded_rows(rows, headers, max_rows)
        return records, _column_metadata(headers, records), worksheet.title
    finally:
        workbook.close()


def parse_sqlite(content: bytes, table_name: str | None, max_rows: int, max_columns: int):
    if not content.startswith(b"SQLite format 3\x00"):
        raise IngestionError("uploaded file is not a SQLite database")
    path = None
    connection = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as handle:
            handle.write(content)
            path = handle.name
        connection = sqlite3.connect(f"file:{Path(path).as_posix()}?mode=ro", uri=True)
        connection.row_factory = sqlite3.Row
        tables = [row[0] for row in connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
        ).fetchall()]
        if not tables:
            raise IngestionError("SQLite database contains no user tables")
        selected = table_name or tables[0]
        if selected not in tables:
            raise IngestionError("requested table does not exist")
        quoted = '"' + selected.replace('"', '""') + '"'
        info = connection.execute(f"PRAGMA table_info({quoted})").fetchall()
        if len(info) > max_columns:
            raise IngestionError(f"source exceeds the {max_columns} column limit")
        count = connection.execute(f"SELECT COUNT(*) FROM {quoted}").fetchone()[0]
        if count > max_rows:
            raise IngestionError(f"source exceeds the {max_rows} row limit")
        rows = connection.execute(f"SELECT * FROM {quoted}").fetchall()
        records = [{key: _json_value(row[key]) for key in row.keys()} for row in rows]
        columns = [{"name": row[1], "declared_type": row[2] or "", "nullable": not bool(row[3]),
                    "primary_key": bool(row[5])} for row in info]
        return records, columns, selected, tables
    except sqlite3.DatabaseError as exc:
        raise IngestionError("invalid or unsupported SQLite database") from exc
    finally:
        if connection is not None:
            connection.close()
        if path:
            try:
                os.unlink(path)
            except FileNotFoundError:
                pass


def _limits() -> tuple[int, int, int]:
    return (int(current_app.config.get("SOURCE_MAX_BYTES", DEFAULT_MAX_BYTES)),
            int(current_app.config.get("SOURCE_MAX_ROWS", DEFAULT_MAX_ROWS)),
            int(current_app.config.get("SOURCE_MAX_COLUMNS", DEFAULT_MAX_COLUMNS)))


def _upload_and_area():
    upload = request.files.get("file")
    if not upload or not upload.filename:
        raise IngestionError("file is required")
    try:
        area_id = int(request.form.get("audit_area_id", ""))
    except ValueError as exc:
        raise IngestionError("valid audit_area_id is required") from exc
    area = db.session.get(AuditArea, area_id)
    if not area:
        raise IngestionError("audit area not found")
    max_bytes, max_rows, max_columns = _limits()
    content = upload.stream.read(max_bytes + 1)
    if len(content) > max_bytes:
        raise IngestionError(f"file exceeds the {max_bytes} byte limit")
    if not content:
        raise IngestionError("file is empty")
    return upload, area, content, max_rows, max_columns


def _save_source(area, upload, source_type, records, columns, extra=None):
    name = request.form.get("name", "").strip() or Path(upload.filename).stem
    if not name:
        raise IngestionError("source name is required")
    config = {"records": records, "columns": columns, "original_filename": Path(upload.filename).name}
    config.update(extra or {})
    source = DataSource(name=name[:128], source_type=source_type, config=config,
                        audit_area=area, last_sync=utcnow())
    db.session.add(source)
    db.session.commit()
    return source


@data_sources_bp.post("/api/data-sources/upload")
def upload_tabular_source():
    try:
        upload, area, content, max_rows, max_columns = _upload_and_area()
        extension = Path(upload.filename).suffix.lower()
        if extension not in ALLOWED_TABULAR_EXTENSIONS:
            raise IngestionError("only CSV and XLSX files are accepted")
        if extension == ".csv":
            records, columns = parse_csv(content, max_rows, max_columns)
            source = _save_source(area, upload, "csv", records, columns)
        else:
            records, columns, sheet = parse_xlsx(content, max_rows, max_columns,
                                                  request.form.get("sheet_name") or None)
            source = _save_source(area, upload, "xlsx", records, columns, {"sheet_name": sheet})
        return jsonify(source_id=source.id, name=source.name, source_type=source.source_type,
                       record_count=len(records), columns=columns, preview=records[:10]), 201
    except IngestionError as exc:
        return jsonify(error=str(exc)), 400


@data_sources_bp.post("/api/data-sources/sqlite")
def upload_sqlite_source():
    try:
        upload, area, content, max_rows, max_columns = _upload_and_area()
        records, columns, table, tables = parse_sqlite(content, request.form.get("table_name") or None,
                                                        max_rows, max_columns)
        source = _save_source(area, upload, "sqlite", records, columns,
                              {"table_name": table, "available_tables": tables})
        return jsonify(source_id=source.id, name=source.name, source_type="sqlite", table_name=table,
                       available_tables=tables, record_count=len(records), columns=columns,
                       preview=records[:10]), 201
    except IngestionError as exc:
        return jsonify(error=str(exc)), 400


@data_sources_bp.get("/api/data-sources/<int:source_id>/schema")
def source_schema(source_id):
    source = db.get_or_404(DataSource, source_id)
    config = source.config or {}
    return jsonify(source_id=source.id, source_type=source.source_type,
                   record_count=len(config.get("records", [])), columns=config.get("columns", []),
                   table_name=config.get("table_name"), sheet_name=config.get("sheet_name"))


@data_sources_bp.get("/api/data-sources/<int:source_id>/preview")
def source_preview(source_id):
    source = db.get_or_404(DataSource, source_id)
    try:
        limit = min(max(int(request.args.get("limit", 10)), 1), 100)
    except ValueError:
        return jsonify(error="limit must be an integer"), 400
    records = (source.config or {}).get("records", [])
    return jsonify(source_id=source.id, total_records=len(records), records=records[:limit])
